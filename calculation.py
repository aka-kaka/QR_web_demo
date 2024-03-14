"""
    NO
"""


from calendar import monthrange
from datetime import datetime

from openpyxl import load_workbook

from solutions import InsertDataBD, curdir, os_sep


class Calculation(InsertDataBD):
    """
    No Doc
    """
    def __init__(
            self,
            file_extension='xls',
            file_name=None,) -> None:

        super().__init__(
            file_extension=file_extension,
            file_name=file_name,)

        self.begin_of_per = 0
        self.perf_gross = 0
        self.end_of_per = 0
        self.date = {}
        self.get_data()
        self.path_input = curdir + os_sep + "input" + os_sep
        self.coef = self.get_coef()

    def __exit__(self, types, value, traceback):
        # print(traceback, value, types)
        self.connections.close()

    def get_coef(self):
        """
        NO DOC
        """
        coef_tmp = dict()
        doc = load_workbook(
            filename=self.path_input+'QR_cashflows.xlsx',
            read_only=True, data_only=True)
        for i in self.get_for_db(
                sql="""SELECT DISTINCT vic.code, vic.min_val, vic.max_val
                FROM View_Ins_Calc as vic
                WHERE diff = ?""",
                task=(1,),
                conn=self.connections):
            tmp = list(doc[f'{i[0]}rate'].iter_rows(
                    min_row=34, max_row=34,
                    min_col=0, values_only=True))[0][-1]
            if tmp < i[1]:
                tmp = i[1]
            if tmp > i[2]:
                tmp = i[2]
            coef_tmp.update({i[0]: tmp})
        return coef_tmp

    def get_suck_fee(self, id_sf):
        """
        no doc
        """

        return self.get_for_db(
            sql="""SELECT percent, bring, percent_sf
            FROM Success_Fee
            WHERE id_success_fee_type=?
            ORDER BY percent;""",
            task=(id_sf,),
            conn=self.connections
        )

    def get_days_of_year(self, year=datetime.now().year) -> int:
        """
        no doc
        """
        if (year % 4 == 0 and year % 100 != 0) or year % 400 == 0:
            res = 366
        else:
            res = 365
        return res

    def calculation_stair(self,
                          perc_year: float,
                          suck_fee: list,) -> float:
        """
        No DOC
        """
        if perc_year <= 0:
            return 0
        result = .0
        for nums, items in enumerate(suck_fee):  # [:: -1]
            if (perc_year // items[0] >= 1.0 and
                    items[1] > 0):
                result += (items[2])
            else:
                if nums > 0:
                    result += (
                        items[0] *
                        (perc_year - suck_fee[nums - 1][0])) / 100
                break
        return result

    def calculation_diff(self,
                         perc_year: float,
                         suck_fee: list,
                         coef: float,) -> float:
        """
        No DOC
        """
        if perc_year <= 0:
            return 0
        result = .0
        len_tab = len(suck_fee) - 1
        for nums, items in enumerate(suck_fee):
            tmp = items[0] * coef
            if perc_year // tmp > 1.0:
                if nums < len_tab:
                    result += coef * items[1]
                else:
                    result += (perc_year - tmp) * items[1]
                    break

            elif perc_year // tmp == 1.0 and (perc_year - tmp > 0):
                result += (perc_year - tmp) * items[1]
        return result

    def get_perc_year(self, date: str, perf_gross):
        """
        no doc
        """
        _, month, year = [int(i) for i in date.split('.')]

        day_fo_mont = monthrange(year=year, month=month)[1]
        perc_year = perf_gross / day_fo_mont * self.get_days_of_year(year)
        return perc_year

    def calculation_new(self, curensy, date, perf_gross):
        """
        NO DOC
        """
        # if perf_gross <= 0:
        #     return 0

        key = self.get_for_db(
            sql="""SELECT key_success_fee_type, diff, key_groups, code
            FROM View_Ins_Calc vic
            WHERE code = ?""",
            task=((curensy,)),
            conn=self.connections)
        res = []
        for item in key:

            perc_year = self.get_perc_year(
                date=date,
                perf_gross=perf_gross,
            )
            if item[1] == 0:

                res.append((
                    self.calculation_stair(
                        suck_fee=self.get_suck_fee(item[0]),
                        perc_year=perc_year),
                    item[2], item[3], perc_year))

            elif item[1] == 1:

                res.append((
                    self.calculation_diff(
                        suck_fee=self.get_suck_fee(item[0]),
                        perc_year=perc_year,
                        coef=self.coef[curensy],),
                    item[2], item[3], perc_year))
        return res

    def calculation_all(self, val_to_pred=None):
        """
        no doc
        """
        result = {}

        if val_to_pred:
            tmp_bool = True
        else:
            return "it is a DEMO"

        for nums, items in enumerate(val_to_pred):
            day_month, aum_bop, success_fee_act = items[0: 3]
            perf_gross, key_period, managment_fee = items[3: 6]
            dates_per = [int(i) for i in items[6].split('-')]
            dates = items[6]
            key_contract = items[8]
            tmp = self.get_days_of_year(dates_per[0])
            profit_month = round(perf_gross * aum_bop / 100, 2)
            aum_change = round((perf_gross * aum_bop / 100) / dates_per[2] *
                               day_month, 2)
            aum_eop_gross = round(aum_bop * (1 + (perf_gross / 100) /
                                             dates_per[2] * day_month), 2)
            max_prev_per = self.get_for_db(
                sql="""SELECT MAX(p.eof_bop) FROM Period as p
                WHERE p.id_contract = ? AND p.dates < ?""",
                task=((key_contract, dates)),
                conn=self.connections)[0][0]
            qr_perf_share = (
                (success_fee_act / tmp * dates_per[2]) / perf_gross)

            # print(f'{max_prev_per=}')
            if max_prev_per is not None:
                if max_prev_per < aum_eop_gross:
                    sf = round(
                        qr_perf_share * (aum_eop_gross - max_prev_per),
                        2)
                else:
                    sf = 0
            else:
                sf = round(qr_perf_share * aum_change, 2)

            avg_dep = round((aum_eop_gross + aum_bop) / 2, 2)
            mf = round(avg_dep * managment_fee / tmp * day_month,
                       2)

            incom = sf + mf
            aum_change_net = round(aum_change - incom, 2)
            aum_eop_net = round(aum_eop_gross - incom, 2)
            if tmp_bool:
                # print(f'{dates=}',
                #       f'{aum_bop=}',
                #       f'{profit_month=}',
                #       f'{aum_change=}',
                #       f'{aum_eop_gross=}',
                #       f'{avg_dep=}',
                #       f'{mf=}',
                #       f'{sf=}',
                #       f'{incom=}',
                #       f'{aum_eop_net=}',
                #       f'{qr_perf_share=}',
                #       sep="\n")
                # print("\n", "*" * 10, "\n")

                result.update(
                    {nums: [
                        key_contract,
                        dates,
                        aum_bop,
                        profit_month,
                        aum_change,
                        aum_eop_gross,
                        avg_dep,
                        mf,
                        sf,
                        incom,
                        aum_eop_net,
                        qr_perf_share,
                        aum_change_net,
                        items[-3],
                        items[-2],
                        items[-1],

                    ]})
        return result
